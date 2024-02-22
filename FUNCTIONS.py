# functions

import itertools
import re
from ast import literal_eval
from collections import Counter
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from scipy.stats import mannwhitneyu, shapiro
import os
import pandas as pd


def clean_text(text):
    # Remove any characters not in the a-zA-ZÀ-ÿ range and punctuation
    text = re.sub(r"[^a-zA-ZÀ-ÿ.\s]", "", text)
    # Remove control characters
    text = re.sub(r"[\x00-\x1f\x7f-\x9f]", "", text)
    text = text.lower()
    return text

def calculate_base_statistics(text):
    # Count total letters, words, sentences, and calculate average word length
    total_letters = len(re.findall(r"\w", text))
    total_words = len(re.findall(r"\w+", text))
    total_sentences = len(re.findall(r"[.!?]", text))
    average_word_length = total_letters / total_words if total_words else 0

    return total_letters, total_words, total_sentences, average_word_length

def process_file(input_file):
    print(f"Processing file: {input_file}")

    with open(input_file, 'r', encoding='utf-8', errors='replace') as file:
        text = file.read()

    text = clean_text(text)

    total_letters, total_words, total_sentences, average_word_length = calculate_base_statistics(text)
    character_count = Counter(text)
    total_count = sum(character_count.values())

    df = pd.DataFrame(list(character_count.items()), columns=['Character', 'Count'])
    df.loc['Total'] = ['Total General', total_count]
    df['Total Letters'] = total_letters
    df['Total Words'] = total_words
    df['Total Sentences'] = total_sentences
    df['Average Word Length'] = average_word_length

    output_path = os.path.splitext(input_file)[0] + '.xlsx'
    df.to_excel(output_path, index=False)
    print(f'Successfully generated spreadsheet: {os.path.abspath(output_path)}')

    min_count = df['Count'].min()
    max_count = df['Count'].max()
    count_variance = df['Count'].var()
    ordered_characters = list(df.sort_values(by='Count')['Character'])
    # Removing "Total General" from the list
    ordered_characters.remove('Total General')
    # Removing spaces from the list
    if ' ' in ordered_characters:
        ordered_characters.remove(' ')

    return total_count, min_count, max_count, count_variance, ordered_characters, total_letters, total_words, total_sentences, average_word_length

def process_folder(folder):
    print(f"Processing folder: {os.path.abspath(folder)}")

    summary = pd.DataFrame(columns=["ID", "folder", "name", "total", "min", "max", "var", "list", "total_letters", "total_words", "total_sentences", "average_word_length"])
    count = 0
    for root, _, files in os.walk(folder):
        print(f"Analyzing directory: {root}")
        for file in files:
            if file.endswith('.txt'):
                count += 1
                file_path = os.path.join(root, file)
                total, min_count, max_count, count_variance, ordered_characters, total_letters, total_words, total_sentences, average_word_length = process_file(file_path)
                summary = pd.concat([summary, pd.DataFrame([{
                    "ID": count,
                    "name": file,
                    "folder": folder,
                    "total": total,
                    "min": min_count,
                    "max": max_count,
                    "var": count_variance,
                    "list": ordered_characters,
                    "total_letters": total_letters,
                    "total_words": total_words,
                    "total_sentences": total_sentences,
                    "average_word_length": average_word_length
                }], index=[0])], ignore_index=True)
            else:
                print(f"Ignored file (not .txt): {file}")

    summary.to_excel(os.path.join(folder, 'summary.xlsx'), index=False)
    print(f'Successfully generated summary: {os.path.abspath(os.path.join(folder, "summary.xlsx"))}')



def aggregate(*folders):

    # List to store each DataFrame
    dataframes = []

    for folder in folders:
        summary_path = os.path.join(folder, 'summary.xlsx')
        if os.path.isfile(summary_path):
            df = pd.read_excel(summary_path)
            dataframes.append(df)
        else:
            print(f"Summary.xlsx file not found in folder: {folder}")

    # Concatenating all DataFrames
    overall_summary = pd.concat(dataframes, ignore_index=True)

    # Saving the aggregated DataFrame as Excel
    overall_summary.to_excel('overall_summary.xlsx', index=False)
    print('Overall summary successfully created: overall_summary.xlsx')





def bubble_sort(list):
    """Performs a bubble sort on the generated normalized sequence.
    This routine executes bubble sort because it's necessary to count the ordering steps of the normalized sequence,
    in order to measure the value of the utilized Kendall Tau metric."""

    elements = len(list) - 1
    sorted = False
    counter = 0
    while not sorted:
        sorted = True
        for i in range(elements):
            if list[i] > list[i + 1]:
                list[i], list[i + 1] = list[i + 1], list[i]
                counter += 1
                sorted = False
    return counter


def ekt(list1, list2):
    """Takes two ordered lists as input.
    The return of the function is the normalized amount of disarray between the two sequences, i.e.,
    the number of movements necessary to revert sequence 2 back to the state of sequence 1, counting
    one unit for each displacement to the left or right. Thus, for an element to change positions
    with a neighbor, it will cost 1.
    The cost to insert an element at position i, the value of i itself, and the cost to
    remove an element from position j, the value of j, and the cost to move an element from position
    i to position j as being j-i are considered. Movements must be carried out with the minimum of
    unit displacements (considering the unit displacement as being the permutation of a single element from position i
    with its neighbor i-1 or i+1). Therefore, a bubble sort algorithm is utilized for arranging the elements,
    and the necessary steps are counted. The normalization denominator (which ensures the metric value between 0 and 1)
    is defined as the maximum number of steps necessary for, given x, y, z the number of common elements,
    excluded and inserted to transform list1 into list2.
    For this function, metric characteristics were tested, and ordered_level1(X,Y) = ekt(Y, X).
    Also, other properties (metric is zero when the base is null) and the triangle inequality are valid.
    The metric values are always in the closed interval from 0 to 1."""

    m = len(list1)
    n = len(list2)
    a = set(list1)
    b = set(list2)
    n1 = len(a.intersection(b))
    term1 = 0
    term2 = 0
    term3 = 0
    for x in range(n1 + 1, n + 1, 1):
        term1 += x
    for x in range(n1 + 1, m + 1, 1):
        term2 += x
    for x in range(n1 - 1, 0, -1):
        term3 += x
    norm = term1 + term2 + term3

    Ynorm = [0] * (len(list2) + 1)
    k = 0
    for i in range(m):
        for j in range(n):
            if list1[i] == list2[j]:
                Ynorm[j + 1] = i + 1
    for j in range(n):
        if list2[j] not in list1:
            k -= 1
            Ynorm[j + 1] = k
    for i in range(m):
        if list1[i] not in list2:
            Ynorm.insert(0, i + 1)
    print(Ynorm)
    z = bubble_sort(Ynorm)
    return z / norm

def common_elements(list1, list2):
    common1 = [element for element in list1 if element in list2]
    common2 = [element for element in list2 if element in list1]
    return common1, common2






def matrix(input_file, output_ekt, output_kt):
    df = pd.read_excel(input_file)
    df['list'] = df['list'].apply(literal_eval)
    pairs = list(itertools.product(df.index, repeat=2))

    ekt_results = pd.DataFrame(index=df.index, columns=df.index)
    kt_results = pd.DataFrame(index=df.index, columns=df.index)

    for pair in pairs:
        id1, folder1, list1 = df.loc[pair[0], ["ID", "folder", "list"]]
        id2, folder2, list2 = df.loc[pair[1], ["ID", "folder", "list"]]

        # Calcula EKT e KT com elementos comuns
        res_ekt = ekt(list1, list2)
        ekt_results.loc[pair[0], pair[1]] = res_ekt

        list1_common, list2_common = common_elements(list1, list2)
        res_kt = ekt(list1_common, list2_common)
        kt_results.loc[pair[0], pair[1]] = res_kt

    # Ajustando índices e colunas para usar o formato 'Pasta-ID'
    ekt_results.index = df.apply(lambda row: f'{row["folder"]}-{row["ID"]}', axis=1)
    ekt_results.columns = ekt_results.index

    kt_results.index = ekt_results.index
    kt_results.columns = ekt_results.index

    # Salvando os resultados
    ekt_results.to_excel(output_ekt, index=True)
    kt_results.to_excel(output_kt, index=True)






def calculate_statistics(input_file, output_file, pairs):
    data = pd.read_excel(input_file, sheet_name='Sheet2')

    # Create a new workbook if it doesn't exist
    if not os.path.isfile(output_file):
        wb = Workbook()
        wb.save(output_file)

    book = load_workbook(output_file)

    # Get the active worksheet
    sheet = book.active

    for pair in pairs:
        n1 = pair[0]
        n_list = pair[1]
        for n in n_list:
            # Remove NaN values
            data1 = data[n1].dropna()
            data2 = data[n].dropna()

            # Mann-Whitney U Test
            result = mannwhitneyu(data1, data2, alternative='two-sided')
            sheet.append([f'Mann-Whitney test for {n1} vs {n} U value: {result.statistic}, p value: {result.pvalue}'])

            # Shapiro-Wilk Test for n1
            shapiro_result_n1 = shapiro(data1)
            sheet.append([f'Shapiro-Wilk test for {n1}: statistic={shapiro_result_n1[0]}, p={shapiro_result_n1[1]}'])

            # Shapiro-Wilk Test for n
            shapiro_result_n = shapiro(data2)
            sheet.append([f'Shapiro-Wilk test for {n}: statistic={shapiro_result_n[0]}, p={shapiro_result_n[1]}'])

            # mean, median, variance for n1
            mean_n1 = np.mean(data1)
            median_n1 = np.median(data1)
            var_n1 = np.var(data1, ddof=1)
            sheet.append([f'Mean for {n1}: {mean_n1}, Median: {median_n1}, Variance: {var_n1}'])

            # mean, median, variance for n
            mean_n = np.mean(data2)
            median_n = np.median(data2)
            var_n = np.var(data2, ddof=1)
            sheet.append([f'Mean for {n}: {mean_n}, Median: {median_n}, Variance: {var_n}'])

            # Cohen's d
            std_dev_n1 = np.sqrt(var_n1)
            std_dev_n = np.sqrt(var_n)
            cohen_d = (mean_n - mean_n1) / ((std_dev_n1 + std_dev_n) / 2)
            sheet.append([f"Cohen's d for {n1} vs {n}: {cohen_d}"])

    # Save the workbook
    book.save(output_file)


pairs = [
    ('ES-ES', ['ES-IT', 'ES-AL', 'ES-IN', 'ES-PT']),
    ('IN-IN', ['IN-AL', 'IN-PT', 'IN-IT', 'ES-IN']),
    ('PT-PT', ['PT-AL', 'PT-IT', 'ES-PT', 'IN-PT']),
    ('IT-IT', ['ES-IT', 'IN-IT', 'IN-PT', 'IT-AL']),
    ('AL-AL', ['IT-AL', 'ES-AL', 'PT-AL', 'IN-AL'])
]
